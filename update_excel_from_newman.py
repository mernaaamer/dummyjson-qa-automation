"""
Updates DummyJSON_QA_Test_Cases.xlsx automatically from a Newman JSON report.

USAGE (run this after every test run):
    python3 update_excel_from_newman.py newman-report.json

WHAT IT DOES:
    - Reads newman-report.json (produced by:
        newman run collection/DummyJSON_API_QA_Collection.postman_collection.json
          -e collection/DummyJSON_QA.postman_environment.json
          --reporters cli,json --reporter-json-export newman-report.json)
    - For every request in the report, finds the matching Test Case row
      (using the same mapping logic as map_testcases_to_requests.py)
    - Fills in: Execution Status (Passed/Failed), Actual Result, Executed By, Execution Date
    - Saves the Excel file in place (keeps a .bak backup of the previous version)

FIX (this version): data rows start at row 3 in every sheet (row 1 = sheet title,
row 2 = headers), not row 5. The previous version skipped the first two test
cases in every sheet because of this.
"""
import sys, json, re, shutil
from datetime import date
import openpyxl

EXCEL_PATH = "DummyJSON_QA_Test_Cases.xlsx"
MAPPING_PATH = "testcase_to_request_mapping.json"  # generated once by map_testcases_to_requests.py


def load_report(path):
    with open(path) as f:
        return json.load(f)


def summarize_execution(execution):
    """Return ('Passed'|'Failed', actual_result_text) for one newman execution entry."""
    assertions = execution.get("assertions", [])
    failed = [a for a in assertions if a.get("error")]
    status_code = execution.get("response", {}).get("code", "N/A")
    if not assertions:
        return "Not Executed", "No assertions ran"
    if failed:
        reasons = "; ".join(f"{a['assertion']}: {a['error'].get('message','')}" for a in failed)
        return "Failed", f"HTTP {status_code}. Failed checks: {reasons}"[:500]
    return "Passed", f"HTTP {status_code}. All {len(assertions)} checks passed."


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 update_excel_from_newman.py <newman-report.json>")
        sys.exit(1)

    report = load_report(sys.argv[1])
    with open(MAPPING_PATH) as f:
        tc_to_request = json.load(f)
    request_to_tc = {v: k for k, v in tc_to_request.items()}

    # Build request-name -> result from the newman report
    results_by_request = {}
    for execution in report.get("run", {}).get("executions", []):
        req_name = execution.get("item", {}).get("name", "")
        status, actual = summarize_execution(execution)
        results_by_request[req_name] = (status, actual)

    shutil.copy(EXCEL_PATH, EXCEL_PATH.replace(".xlsx", ".bak.xlsx"))
    wb = openpyxl.load_workbook(EXCEL_PATH)

    updated, skipped = 0, 0
    unmatched_requests = []
    today = date.today().isoformat()

    for sheet_name in wb.sheetnames:
        if sheet_name == "Legend & Overview":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=3):  # FIX: was min_row=5
            tc_id_cell = row[0]
            tc_id = tc_id_cell.value
            if not tc_id or tc_id not in tc_to_request:
                continue
            req_name = tc_to_request[tc_id]
            if req_name not in results_by_request:
                skipped += 1
                unmatched_requests.append((tc_id, req_name))
                continue
            status, actual = results_by_request[req_name]
            row[8].value = status          # Execution Status
            row[9].value = actual          # Actual Result
            row[11].value = "Newman (automated)"  # Executed By
            row[12].value = today          # Execution Date
            updated += 1

    wb.save(EXCEL_PATH)
    print(f"Updated {updated} test case rows.")
    print(f"Skipped {skipped} (mapped but not found in this report run).")
    if unmatched_requests:
        print("Skipped test cases (Test Case ID -> expected request name):")
        for tc_id, req_name in unmatched_requests:
            print(f"  - {tc_id} -> '{req_name}'")
    print(f"Backup saved as {EXCEL_PATH.replace('.xlsx', '.bak.xlsx')}")


if __name__ == "__main__":
    main()
